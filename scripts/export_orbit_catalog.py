import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = REPO_ROOT / "cases" / "orbit_catalog.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "cases" / "orbit_catalog.csv"


def row_has_value(row):
    return any(value not in (None, "") for value in row)


def normalize_cell(value):
    if value is None:
        return ""
    return value


def export_orbit_catalog(input_path, output_path, sheet_name):
    workbook = load_workbook(input_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = [
        [normalize_cell(cell.value) for cell in row]
        for row in worksheet.iter_rows()
    ]
    rows = [row for row in rows if row_has_value(row)]

    if not rows:
        raise ValueError(f"No rows found in {input_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    return len(rows) - 1


def parse_args():
    parser = argparse.ArgumentParser(
        description="Export the human-edited orbit catalog Excel file to CSV."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--sheet", default="orbit_catalog")
    return parser.parse_args()


def main():
    args = parse_args()
    orbit_count = export_orbit_catalog(args.input, args.output, args.sheet)
    print(f"Exported {orbit_count} orbits")
    print(f"Input : {args.input}")
    print(f"Output: {args.output}")


if __name__ == "__main__":
    main()
